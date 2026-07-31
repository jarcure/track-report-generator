"""
Reads the Landmark Surveys CSV export, splits it into rounds (one per
unique Event Time), and fills a copy of the template for each round
using the editable point_mapping.csv config file.

- Point Name found in the mapping -> writes StdDevNorthing/Easting/Elevation
  into that TMP's DN/DE/DELV cells.
- Point Name not found in that round's data -> cell stays blank.
- Each round is saved as its own file, named by date and time.

Usage: python3 process.py <csv_file> <template_file> <mapping_file> <output_dir> [--limit N]
"""
import csv
import sys
import os
import shutil
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import LineChart, BarChart, Reference

NULL_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

def build_summary_workbook(summary_rows, out_path):
    """Builds a summary workbook with a data table and two native Excel charts:
    a trend line of missing points per round, and a bar chart ranking which
    points go missing most often."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = ["Date", "Time", "Points Found", "Points Missing", "Missing Point Names", "Filename"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for r, row in enumerate(summary_rows, start=2):
        ws.cell(row=r, column=1, value=row["Date"])
        ws.cell(row=r, column=2, value=row["Time"])
        ws.cell(row=r, column=3, value=row["Points Found"])
        ws.cell(row=r, column=4, value=row["Points Missing"])
        ws.cell(row=r, column=5, value=row["Missing Point Names"])
        ws.cell(row=r, column=6, value=row["Filename"])

    last_row = len(summary_rows) + 1

    # Trend chart: Points Missing per round
    trend = LineChart()
    trend.title = "Points missing per round"
    trend.y_axis.title = "Points missing"
    trend.x_axis.title = "Round"
    data = Reference(ws, min_col=4, min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=6, min_row=2, max_row=last_row)
    trend.add_data(data, titles_from_data=True)
    trend.set_categories(cats)
    trend.height, trend.width = 8, 18
    ws.add_chart(trend, f"H1")

    # Ranking sheet: which points go missing most often
    counts = {}
    for row in summary_rows:
        names = row["Missing Point Names"]
        if names:
            for name in names.split(", "):
                if name:
                    counts[name] = counts.get(name, 0) + 1

    ws2 = wb.create_sheet("Missing Point Ranking")
    ws2.cell(row=1, column=1, value="Point Name").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Times Missing").font = Font(bold=True)
    for r, (name, count) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), start=2):
        ws2.cell(row=r, column=1, value=name)
        ws2.cell(row=r, column=2, value=count)

    if counts:
        bar = BarChart()
        bar.title = "Which points go missing most often"
        bar.y_axis.title = "Times missing"
        last_r = len(counts) + 1
        data = Reference(ws2, min_col=2, min_row=1, max_row=last_r)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=last_r)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height, bar.width = 8, 18
        ws2.add_chart(bar, "D1")

    wb.save(out_path)

def load_mapping(path):
    mapping = {}
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            mapping[row["PointName"]] = row
    return mapping

def load_rounds(csv_path):
    rounds = defaultdict(list)
    with open(csv_path, newline="") as f:
        for row in csv.DictReader(f):
            ts = row.get("Event Time (Eastern Standard Time)")
            if ts:
                rounds[ts].append(row)
    return rounds

LS_FILE_CELLS = ["I1", "I44", "I87", "I130"]
PROJECT_CELLS = ["D7", "D50", "D93", "D136"]
CONTRACTOR_CELLS = ["D8", "D51", "D94", "D137"]
CERT_CELLS = ["H5", "H48", "H91", "H134"]
REV_CELLS = ["I8", "I51", "I94", "I137"]

def safe_write(wb, sheet_name, cell_ref, value, fill=None):
    if not isinstance(sheet_name, str) or not sheet_name.strip():
        sheet_name = "Sheet1"
    else:
        sheet_name = sheet_name.strip()
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]
    ws[cell_ref] = value
    if fill:
        ws[cell_ref].fill = fill
    return True

def fill_template(template_path, out_path, rows, mapping, ls_file="25146",
                   project_name="STAFFORD ROAD SANITARY SEWER PHASE I",
                   contractor_name="MONTANA CONSTRUCTION",
                   cert_number="24GA2872300", rev_number="0"):
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Sheet1"]

    ts = rows[0]["Event Time (Eastern Standard Time)"]
    dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
    date_str = dt.strftime("%m-%d-%y")
    time_str = dt.strftime("%I:%M%p").lstrip("0").lower()

    # Header blocks repeat at these rows for each of the 4 stacked tracks
    for date_row, time_row in [(2, 3), (45, 46), (88, 89), (131, 132)]:
        ws[f"I{date_row}"] = date_str
        ws[f"I{time_row}"] = time_str
    for cell_ref in LS_FILE_CELLS:
        ws[cell_ref] = ls_file
    for cell_ref in PROJECT_CELLS:
        ws[cell_ref] = project_name
    for cell_ref in CONTRACTOR_CELLS:
        ws[cell_ref] = contractor_name
    for cell_ref in CERT_CELLS:
        ws[cell_ref] = f"Cert. of Auth. #{cert_number}"
    for cell_ref in REV_CELLS:
        ws[cell_ref] = f"REV {rev_number}"

    matched, unmatched = [], []
    for row in rows:
        name = row["Point Name"]
        cfg = mapping.get(name)
        if not cfg:
            unmatched.append(name)
            continue
        sheet = cfg.get("Sheet", "Sheet1")
        safe_write(wb, sheet, cfg["Cell_DN"], float(row["StdDevNorthing"]))
        safe_write(wb, sheet, cfg["Cell_DE"], float(row["StdDevEasting"]))
        safe_write(wb, sheet, cfg["Cell_DELV"], float(row["StdDevElevation"]))
        matched.append(name)

    # Any TMP slot in the mapping that had no matching row this round gets NULL,
    # highlighted yellow so it's obvious at a glance
    found_names = {row["Point Name"] for row in rows}
    missing_points = []
    for name, cfg in mapping.items():
        if name not in found_names:
            sheet = cfg.get("Sheet", "Sheet1")
            safe_write(wb, sheet, cfg["Cell_DN"], "NULL", NULL_FILL)
            safe_write(wb, sheet, cfg["Cell_DE"], "NULL", NULL_FILL)
            safe_write(wb, sheet, cfg["Cell_DELV"], "NULL", NULL_FILL)
            missing_points.append(name)

    wb.save(out_path)
    return date_str, time_str, matched, unmatched, missing_points

def main():
    csv_path, template_path, mapping_path, out_dir = sys.argv[1:5]
    limit = None
    if "--limit" in sys.argv:
        limit = int(sys.argv[sys.argv.index("--limit") + 1])

    os.makedirs(out_dir, exist_ok=True)
    mapping = load_mapping(mapping_path)
    rounds = load_rounds(csv_path)

    print(f"Found {len(rounds)} rounds total.")
    items = list(rounds.items())
    if limit:
        items = items[:limit]

    summary_rows = []
    for ts, rows in items:
        dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        fname = dt.strftime("%m-%d-%y_%I%M%p").lstrip("0").lower() + ".xlsx"
        out_path = os.path.join(out_dir, fname)
        date_str, time_str, matched, unmatched, missing_points = fill_template(
            template_path, out_path, rows, mapping
        )
        print(f"{fname}: matched {len(matched)}, unmatched {len(unmatched)} {unmatched if unmatched else ''}")
        summary_rows.append({
            "Date": date_str, "Time": time_str,
            "Points Found": len(matched), "Points Missing": len(missing_points),
            "Missing Point Names": ", ".join(missing_points),
            "Filename": fname,
        })

    summary_path = os.path.join(out_dir, "_summary_log.csv")
    with open(summary_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Date", "Time", "Points Found", "Points Missing", "Missing Point Names", "Filename"
        ])
        writer.writeheader()
        writer.writerows(summary_rows)
    print(f"\nSummary log (CSV) written to {summary_path}")

    summary_xlsx_path = os.path.join(out_dir, "_summary_report.xlsx")
    build_summary_workbook(summary_rows, summary_xlsx_path)
    print(f"Summary report with charts written to {summary_xlsx_path}")

if __name__ == "__main__":
    main()

import streamlit as st
import pandas as pd
import openpyxl
import io
import zipfile
import csv
from collections import defaultdict
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import LineChart, BarChart, Reference

NULL_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

def build_summary_workbook(summary_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    headers = ["Date", "Time", "Points Found", "Points Missing", "Missing Point Names", "Filename"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    for r, row in enumerate(summary_rows, start=2):
        ws.cell(row=r, column=1, value=row["Date"])
        ws.cell(row=r, column=2, value=row["Time"])
        ws.cell(row=r, column=3, value=row["Points Found"])
        ws.cell(row=r, column=4, value=row["Points Missing"])
        ws.cell(row=r, column=5, value=row["Missing Point Names"])
        ws.cell(row=r, column=6, value=row["Filename"])

    last_row = len(summary_rows) + 1
    trend = LineChart()
    trend.title = "Points missing per round"
    trend.y_axis.title = "Points missing"
    data = Reference(ws, min_col=4, min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=6, min_row=2, max_row=last_row)
    trend.add_data(data, titles_from_data=True)
    trend.set_categories(cats)
    trend.height, trend.width = 8, 18
    ws.add_chart(trend, "H1")

    counts = {}
    for row in summary_rows:
        names = row["Missing Point Names"]
        if names:
            for name in names.split(", "):
                if name:
                    counts[name] = counts.get(name, 0) + 1

    ws2 = wb.create_sheet("Missing Point Ranking")
    ws2.cell(row=1, column=1, value="Point Name").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Times Missing").font = Font(bold=True)
    for r, (name, count) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), start=2):
        ws2.cell(row=r, column=1, value=name)
        ws2.cell(row=r, column=2, value=count)
    if counts:
        bar = BarChart()
        bar.title = "Which points go missing most often"
        bar.y_axis.title = "Times missing"
        last_r = len(counts) + 1
        data = Reference(ws2, min_col=2, min_row=1, max_row=last_r)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=last_r)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height, bar.width = 8, 18
        ws2.add_chart(bar, "D1")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

st.set_page_config(page_title="Landmark Surveys - Report Generator", layout="wide")
st.title("Track Monitoring Report Generator")
st.caption("Upload a CSV export, review the point mapping, and download the filled reports.")

TEMPLATE_PATH = "TRACK_MONITORING_REPORT.xlsx"
DEFAULT_MAPPING_PATH = "point_mapping.csv"
HEADER_ROWS = [(2, 3), (45, 46), (88, 89), (131, 132)]
LS_FILE_CELLS = ["I1", "I44", "I87", "I130"]
PROJECT_CELLS = ["D7", "D50", "D93", "D136"]
CONTRACTOR_CELLS = ["D8", "D51", "D94", "D137"]
CERT_CELLS = ["H5", "H48", "H91", "H134"]
REV_CELLS = ["I8", "I51", "I94", "I137"]

st.subheader("1. Report details")
st.write("These update every track section in the template automatically.")
col_a, col_b, col_c = st.columns(3)
with col_a:
    ls_file = st.text_input("LS File #", value="25146")
with col_b:
    project_name = st.text_input("Project", value="STAFFORD ROAD SANITARY SEWER PHASE I")
with col_c:
    contractor_name = st.text_input("Contractor", value="MONTANA CONSTRUCTION")

col_d, col_e = st.columns(2)
with col_d:
    cert_number = st.text_input("Cert. of Auth. #", value="24GA2872300")
with col_e:
    rev_number = st.text_input("REV", value="0")

st.divider()

# ---------- Mapping table (editable in-browser) ----------
if "mapping_df" not in st.session_state:
    st.session_state.mapping_df = pd.read_csv(DEFAULT_MAPPING_PATH)

st.subheader("2. Point mapping")
st.write("Add, remove, or edit rows below. Each Point Name needs a TMP number and the exact cells to write into.")
edited = st.data_editor(
    st.session_state.mapping_df,
    num_rows="dynamic",
    use_container_width=True,
    key="mapping_editor",
)
st.session_state.mapping_df = edited

col1, col2 = st.columns(2)
with col1:
    st.download_button(
        "Download current mapping as CSV",
        edited.to_csv(index=False),
        file_name="point_mapping.csv",
        mime="text/csv",
    )
with col2:
    uploaded_mapping = st.file_uploader("Or upload a saved mapping CSV", type="csv", key="mapping_upload")
    if uploaded_mapping:
        st.session_state.mapping_df = pd.read_csv(uploaded_mapping)
        st.rerun()

st.divider()

# ---------- CSV upload and processing ----------
st.subheader("3. Upload CSV export")
csv_file = st.file_uploader("CSV file from the field device", type="csv", key="data_csv")

def load_rounds(file):
    text = io.TextIOWrapper(file, encoding="utf-8")
    rounds = defaultdict(list)
    for row in csv.DictReader(text):
        ts = row.get("Event Time (Eastern Standard Time)")
        if ts:
            rounds[ts].append(row)
    return rounds

def safe_write(wb, sheet_name, cell_ref, value, fill=None):
    if not isinstance(cell_ref, str) or not CELL_PATTERN.match(cell_ref.strip()):
        return False
    if not isinstance(sheet_name, str) or not sheet_name.strip():
        sheet_name = "Sheet1"
    else:
        sheet_name = sheet_name.strip()
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]
    ws[cell_ref.strip()] = value
    if fill:
        ws[cell_ref.strip()].fill = fill
    return True

def fill_template(rows, mapping, ls_file, project_name, contractor_name, cert_number, rev_number):
    with open(TEMPLATE_PATH, "rb") as f:
        wb = openpyxl.load_workbook(f)
    ws = wb["Sheet1"]

    ts = rows[0]["Event Time (Eastern Standard Time)"]
    dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
    date_str = dt.strftime("%m-%d-%y")
    time_str = dt.strftime("%I:%M%p").lstrip("0").lower()

    for date_row, time_row in HEADER_ROWS:
        ws[f"I{date_row}"] = date_str
        ws[f"I{time_row}"] = time_str

    for cell_ref in LS_FILE_CELLS:
        ws[cell_ref] = ls_file
    for cell_ref in PROJECT_CELLS:
        ws[cell_ref] = project_name
    for cell_ref in CONTRACTOR_CELLS:
        ws[cell_ref] = contractor_name
    for cell_ref in CERT_CELLS:
        ws[cell_ref] = f"Cert. of Auth. #{cert_number}"
    for cell_ref in REV_CELLS:
        ws[cell_ref] = f"REV {rev_number}"

    matched, unmatched = 0, 0
    for row in rows:
        name = row["Point Name"]
        cfg = mapping.get(name)
        if not cfg:
            unmatched += 1
            continue
        sheet = cfg.get("Sheet", "Sheet1")
        ok = (safe_write(wb, sheet, cfg["Cell_DN"], float(row["StdDevNorthing"]))
              and safe_write(wb, sheet, cfg["Cell_DE"], float(row["StdDevEasting"]))
              and safe_write(wb, sheet, cfg["Cell_DELV"], float(row["StdDevElevation"])))
        if ok:
            matched += 1
        else:
            unmatched += 1

    # Any mapped point with no data this round gets NULL, highlighted yellow
    found_names = {row["Point Name"] for row in rows}
    missing_points = []
    for name, cfg in mapping.items():
        if name not in found_names:
            sheet = cfg.get("Sheet", "Sheet1")
            safe_write(wb, sheet, cfg["Cell_DN"], "NULL", NULL_FILL)
            safe_write(wb, sheet, cfg["Cell_DE"], "NULL", NULL_FILL)
            safe_write(wb, sheet, cfg["Cell_DELV"], "NULL", NULL_FILL)
            missing_points.append(name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = dt.strftime("%m-%d-%y_%I%M%p").lstrip("0").lower() + ".xlsx"
    return fname, buf, matched, unmatched, date_str, time_str, missing_points

import re

CELL_PATTERN = re.compile(r"^[A-Z]+[0-9]+$")

def validate_mapping(mapping_df):
    """Returns (clean_lookup_dict, list_of_problem_rows)."""
    clean = {}
    problems = []
    for row in mapping_df.to_dict("records"):
        name = str(row.get("PointName", "")).strip()
        cells = [row.get("Cell_DN"), row.get("Cell_DE"), row.get("Cell_DELV")]
        if not name or name.lower() == "nan":
            continue  # blank row, ignore silently
        bad_cells = [c for c in cells if not isinstance(c, str) or not CELL_PATTERN.match(str(c).strip())]
        if bad_cells:
            problems.append(name)
            continue
        row["Cell_DN"], row["Cell_DE"], row["Cell_DELV"] = [str(c).strip() for c in cells]
        clean[name] = row
    return clean, problems

if csv_file is not None:
    mapping_lookup, problem_rows = validate_mapping(st.session_state.mapping_df)
    if problem_rows:
        st.warning(
            f"These mapping rows are missing a valid cell reference (like C15) and will be "
            f"skipped: {', '.join(problem_rows)}. Fix them in the table above and re-upload if needed."
        )
    rounds = load_rounds(csv_file)
    st.write(f"Found **{len(rounds)}** rounds in this file.")

    max_preview = min(len(rounds), 50)
    limit = st.number_input(
        "How many rounds to process (leave as full count for the whole file)",
        min_value=1, max_value=len(rounds), value=len(rounds),
    )

    if st.button("Generate reports", type="primary"):
        items = list(rounds.items())[:limit]
        zip_buf = io.BytesIO()
        total_matched, total_unmatched = 0, 0
        summary_rows = []
        progress = st.progress(0)
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, (ts, rows) in enumerate(items):
                fname, filebuf, matched, unmatched, date_str, time_str, missing_points = fill_template(
                    rows, mapping_lookup, ls_file, project_name, contractor_name, cert_number, rev_number
                )
                zf.writestr(fname, filebuf.read())
                total_matched += matched
                total_unmatched += unmatched
                summary_rows.append({
                    "Date": date_str, "Time": time_str,
                    "Points Found": matched, "Points Missing": unmatched,
                    "Missing Point Names": ", ".join(missing_points),
                    "Filename": fname,
                })
                progress.progress((i + 1) / len(items))

            summary_df = pd.DataFrame(summary_rows)
            zf.writestr("_summary_log.csv", summary_df.to_csv(index=False))
            summary_xlsx_buf = build_summary_workbook(summary_rows)
            zf.writestr("_summary_report.xlsx", summary_xlsx_buf.read())
        zip_buf.seek(0)

        st.success(f"Generated {len(items)} report(s). {total_matched} point-values written, {total_unmatched} marked NULL (not found that round).")

        st.subheader("4. What this tells you")

        col1, col2 = st.columns(2)
        with col1:
            st.write("**Missing points over time**")
            st.caption("A rising trend usually means an equipment or access issue, not random noise.")
            trend = summary_df[["Filename", "Points Missing"]].copy()
            trend = trend.set_index("Filename")
            st.line_chart(trend)

        with col2:
            st.write("**Which points go missing most often**")
            st.caption("If one point is missing constantly, that's worth checking on site.")
            all_missing = summary_df["Missing Point Names"].dropna()
            counts = {}
            for entry in all_missing:
                for name in str(entry).split(", "):
                    if name:
                        counts[name] = counts.get(name, 0) + 1
            if counts:
                counts_df = pd.DataFrame(
                    sorted(counts.items(), key=lambda x: -x[1]),
                    columns=["Point", "Times missing"]
                ).set_index("Point")
                st.bar_chart(counts_df)
            else:
                st.write("No missing points in this batch, nice.")

        st.subheader("5. Full log")
        st.dataframe(summary_df, use_container_width=True)
        st.download_button(
            "Download all reports (.zip)",
            zip_buf,
            file_name="track_reports.zip",
            mime="application/zip",
        )
else:
    st.info("Upload a CSV to get started.")
